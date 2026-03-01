from pathlib import Path
import shutil

import pandas as pd
from openpyxl import load_workbook

ACTUALS = Path('/Users/josephmoussa/Desktop/mock_structure/2026/atil/fy26_actuals/january_2026/actuals_2026_data.xlsx')
BUDGET = Path('/Users/josephmoussa/Desktop/mock_structure/2026/atil/fy26_budget/budget_mock_data_2026.xlsx')
LOOKUP_XLS = Path('/Users/josephmoussa/Desktop/Cost Centers/Project Codes Consolidation/ATIL Project Codes.xls')


def rename_xlsx_sheet(path: Path, target: str) -> None:
    if not path.exists():
        print(f'MISSING: {path}')
        return
    wb = load_workbook(path)
    if target in wb.sheetnames:
        print(f'OK (already): {path} -> {target}')
        return
    first = wb.sheetnames[0]
    wb[first].title = target
    wb.save(path)
    print(f'RENAMED: {path} [{first} -> {target}]')


def rename_xls_or_fallback(path: Path, target: str) -> None:
    if not path.exists():
        print(f'MISSING: {path}')
        return
    xl = pd.ExcelFile(path)
    if target in xl.sheet_names:
        print(f'OK (already): {path} -> {target}')
        return
    df = pd.read_excel(path, sheet_name=0)
    try:
        import xlwt  # noqa: F401

        tmp = path.with_suffix('.tmp.xls')
        with pd.ExcelWriter(tmp, engine='xlwt') as writer:
            df.to_excel(writer, index=False, sheet_name=target)
        shutil.move(str(tmp), str(path))
        print(f'RENAMED: {path} [first sheet -> {target}]')
    except Exception as exc:
        xlsx_path = path.with_suffix('.xlsx')
        with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name=target)
        print(f'FALLBACK: could not rewrite .xls ({exc})')
        print(f'CREATED: {xlsx_path} with sheet {target}')


if __name__ == '__main__':
    rename_xlsx_sheet(ACTUALS, 'OS extract')
    rename_xlsx_sheet(BUDGET, 'Database')
    rename_xls_or_fallback(LOOKUP_XLS, 'Reference')
